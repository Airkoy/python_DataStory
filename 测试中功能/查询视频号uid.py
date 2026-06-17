import requests
import json
import openpyxl
import os
from datetime import datetime

HEADERS = {
    "Authorization": "eyJhbGciOiJIUzUxMiJ9.eyJzdWIiOiJoZXJtZXNfYWRtaW5AZGF0YXN0b3J5LmNvbS5jbiIsImF1ZGllbmNlIjoid2ViIiwiY3JlYXRlZCI6MTc4MTU4MTI0MzAzMSwicHciOiJoU204VzJNbXVIeT9KQCFOek0yVCIsImV4cCI6MTc4MjE4NjA0M30.e1VOeZahPGrZtN1N7x29xDyr_r9VJuzauMTU1H5pqwIuJBU7x9YyYnXEsuP3FWpQ_P7oG1YNm4wPSJMtc-V0FQ",
    "Content-Type": "application/json"
}

SEARCH_URL = "https://dc.datastory.com.cn/monitor/account/wechatVideo/batchSearch?uuu=banyan"
OUTPUT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "视频号uid查询结果.xlsx")


def load_existing_results():
    """加载已有结果，支持断点续传"""
    if not os.path.exists(OUTPUT_FILE):
        return {}, set()
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb.active
    results = {}
    done = set()
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        name, uid = row
        if name:
            results[str(name).strip()] = str(uid).strip() if uid else ""
            done.add(str(name).strip())
    return results, done


def save_results(results):
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.cell(row=1, column=1, value="账号名称")
    out_ws.cell(row=1, column=2, value="uid")
    for j, (name, uid) in enumerate(results.items(), 2):
        out_ws.cell(row=j, column=1, value=name)
        out_ws.cell(row=j, column=2, value=uid)
    out_wb.save(OUTPUT_FILE)


def search_uid(name):
    data = {"name": [name]}
    try:
        resp = requests.post(SEARCH_URL, headers=HEADERS, data=json.dumps(data), timeout=10)
        result = resp.json()
        uids = result.get("data", {}).get("uids", [])
        if uids:
            return uids[0].get("uid", "")
        return ""
    except Exception as e:
        print(f"[ERROR] {name}: {e}", flush=True)
        return ""


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_file = os.path.join(script_dir, "..", "awt-添加视频号监控.xlsx")

    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    names = [str(row[0]).strip() for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0] and str(row[0]).strip()]
    unique_names = list(dict.fromkeys(names))

    results, done = load_existing_results()
    remaining = [n for n in unique_names if n not in done]
    print(f"共 {len(unique_names)} 个去重账号，已完成 {len(done)}，剩余 {len(remaining)}", flush=True)

    for i, name in enumerate(remaining, 1):
        uid = search_uid(name)
        results[name] = uid
        print(f"[{i}/{len(remaining)}] {name} -> {uid}", flush=True)
        if i % 10 == 0:
            save_results(results)

    save_results(results)
    found = sum(1 for v in results.values() if v)
    print(f"\n结果已保存至: {OUTPUT_FILE}", flush=True)
    print(f"查询成功: {found}/{len(results)}", flush=True)


if __name__ == "__main__":
    main()
