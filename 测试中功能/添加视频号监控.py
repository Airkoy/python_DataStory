import requests
import json
import openpyxl
import os
from datetime import datetime

HEADERS = {
    "Authorization": "eyJhbGciOiJIUzUxMiJ9.eyJzdWIiOiJoZXJtZXNfYWRtaW5AZGF0YXN0b3J5LmNvbS5jbiIsImF1ZGllbmNlIjoid2ViIiwiY3JlYXRlZCI6MTc4MTU4MTI0MzAzMSwicHciOiJoU204VzJNbXVIeT9KQCFOek0yVCIsImV4cCI6MTc4MjE4NjA0M30.e1VOeZahPGrZtN1N7x29xDyr_r9VJuzauMTU1H5pqwIuJBU7x9YyYnXEsuP3FWpQ_P7oG1YNm4wPSJMtc-V0FQ",
    "Content-Type": "application/json"
}

ADD_URL = "https://dc.datastory.com.cn/monitor/account/wechatVideo/add?uuu=banyan"

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE = os.path.join(SCRIPT_DIR, "..", "视频号uid查询结果.xlsx")
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "..", "视频号添加监控结果.xlsx")


def load_existing():
    if not os.path.exists(OUTPUT_FILE):
        return {}, set()
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb.active
    results = {}
    done = set()
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        name = str(row[0]).strip() if row[0] else ""
        uid = str(row[1]).strip() if row[1] else ""
        msg = str(row[2]).strip() if row[2] else ""
        if name:
            results[name] = (uid, msg)
            done.add(name)
    return results, done


def save_results(results):
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.cell(row=1, column=1, value="账号名称")
    out_ws.cell(row=1, column=2, value="uid")
    out_ws.cell(row=1, column=3, value="运行结果")
    for j, (name, (uid, msg)) in enumerate(results.items(), 2):
        out_ws.cell(row=j, column=1, value=name)
        out_ws.cell(row=j, column=2, value=uid)
        out_ws.cell(row=j, column=3, value=msg)
    out_wb.save(OUTPUT_FILE)


def add_monitor(name, uid):
    payload = json.dumps({"accounts": [{"name": name, "uid": uid}]}, ensure_ascii=False)
    try:
        resp = requests.post(ADD_URL, headers=HEADERS, data=payload, timeout=30)
        return resp.text
    except Exception as e:
        return f"ERROR: {e}"


def main():
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active
    accounts = []
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        name = str(row[0]).strip() if row[0] else ""
        uid = str(row[1]).strip() if row[1] else ""
        if name and uid:
            accounts.append((name, uid))

    results, done = load_existing()
    remaining = [(n, u) for n, u in accounts if n not in done]
    print(f"共 {len(accounts)} 个账号，已完成 {len(done)}，剩余 {len(remaining)}", flush=True)

    for i, (name, uid) in enumerate(remaining, 1):
        msg = add_monitor(name, uid)
        results[name] = (uid, msg)
        print(f"[{i}/{len(remaining)}] {name} -> {msg[:80]}", flush=True)
        if i % 10 == 0:
            save_results(results)

    save_results(results)
    ok = sum(1 for _, (_, msg) in results.items() if '"success":true' in msg or '"code":0' in msg)
    print(f"\n结果已保存至: {OUTPUT_FILE}", flush=True)
    print(f"添加成功: {ok}/{len(results)}", flush=True)


if __name__ == "__main__":
    main()
